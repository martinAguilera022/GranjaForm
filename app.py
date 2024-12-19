from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import load_workbook
from pyexcelerate import Workbook
import tempfile
import os
import xlsx2pdf

app = Flask(__name__)

# Ruta para mostrar el formulario


@app.route('/')
def form():
    return render_template('index.html')

# Ruta para procesar los datos del formulario


@app.route('/guardar_datos', methods=['POST'])
def guardar_datos():
    # Extrae los datos enviados en formato JSON desde la solicitud
    
        datos = request.json
        print(datos)
        # Ruta del archivo de plantilla de Excel
        archivo_plantilla = 'plantilla.xlsx'
        if not os.path.exists(archivo_plantilla):
            return jsonify({"status": "error", "message": "El archivo no existe"}), 404
        # Carga el archivo Excel (plantilla existente)
        wb = load_workbook(archivo_plantilla)
        ws = wb.active  # Obtiene la hoja activa del archivo Excel
        print(ws)
        # Escribe los datos enviados por el cliente en celdas específicas del archivo Excel
        ws['B8'] = datos['fecha']  # Fecha
        ws['H8'] = datos['hora']  # Hora
        ws['B10'] = datos['nombreGranja']  # Nombre de la granja
        ws['B11'] = datos['direccion']  # Dirección de la granja
        ws['B12'] = datos['propietario']  # Propietario de la granja

        # Continúa escribiendo datos en las celdas correspondientes según el JSON recibido
        ws['B14'] = datos['tipoPollo']
        ws['B16'] = datos['destino']
        ws['B17'] = datos['cuadrilla1']
        ws['C17'] = datos['cuadrilla2']
        ws['D17'] = datos['cuadrilla3']
        ws['E17'] = datos['cuadrilla4']
        ws['B18'] = datos['cantOperarios1']
        ws['C18'] = datos['cantOperarios2']
        ws['D18'] = datos['cantOperarios3']
        ws['E18'] = datos['cantOperarios4']
        ws['B19'] = datos['mortandad']

        # Información sobre las personas presentes
        ws['E20'] = datos['propietarioPresente']
        ws['G20'] = datos['encargadoPresente']
        ws['I20'] = datos['ningunoPresente']

        # Estado de los caminos
        ws['E22'] = datos['tierra']
        ws['G22'] = datos['asfalto']
        ws['I22'] = datos['mejorado']

        # Datos sobre condiciones de caminos internos
        ws['D25'] = datos['caminosInternosB']
        ws['F25'] = datos['caminosInternosR']
        ws['G25'] = datos['caminosInternosM']
        ws['H25'] = datos['caminosInternosObservaciones']

        # Más datos sobre caminos
        ws['D26'] = datos['caminosHastaB']
        ws['F26'] = datos['caminosHastaR']
        ws['G26'] = datos['caminosHastaM']
        ws['H26'] = datos['caminosHastaObservaciones']

        # Datos sobre galpones y jaulas
        ws['D27'] = datos['tejidosGalponB']
        ws['F27'] = datos['tejidosGalponR']
        ws['G27'] = datos['tejidosGalponM']
        ws['H27'] = datos['tejidosGalponObservaciones']
        ws['D28'] = datos['camaGalponB']
        ws['F28'] = datos['camaGalponR']
        ws['G28'] = datos['camaGalponM']
        ws['H28'] = datos['camaGalponObservaciones']
        ws['D29'] = datos['estadoJaulasB']
        ws['F29'] = datos['estadoJaulasR']
        ws['G29'] = datos['estadoJaulasM']
        ws['H29'] = datos['estadoJaulasObservaciones']

        # Horarios y detalles adicionales
        ws['F32'] = datos['comienzoCarga']
        ws['B32'] = datos['corteAlimento']
        ws['B33'] = datos['horaAyuno']

        # Información sobre el equipo de carga
        ws['B38'] = datos['equipoCarga']

        # Observaciones específicas en áreas clave
        ws['D41'] = datos['manipulacionCargaB']
        ws['F41'] = datos['manipulacionCargaR']
        ws['G41'] = datos['manipulacionCargaM']
        ws['H41'] = datos['manipulacionCargaObservaciones']
        ws['D42'] = datos['encerradoAvesB']
        ws['F42'] = datos['encerradoAvesR']
        ws['G42'] = datos['encerradoAvesM']
        ws['H42'] = datos['encerradoAvesObservaciones']
        ws['D43'] = datos['cargaJaulasB']
        ws['F43'] = datos['cargaJaulasR']
        ws['G43'] = datos['cargaJaulasM']
        ws['H43'] = datos['cargaJaulasObservaciones']
        ws['D44'] = datos['avesMuertasRetiradasB']
        ws['F44'] = datos['avesMuertasRetiradasR']
        ws['G44'] = datos['avesMuertasRetiradasM']
        ws['H44'] = datos['avesMuertasRetiradasObservaciones']

        # Indicadores de bienestar animal
        ws['D48'] = datos['cargaAvesMuertasSi']
        ws['F48'] = datos['cargaAvesMuertasNo']
        ws['G48'] = datos['cargaAvesMuertasNa']
        ws['D49'] = datos['sacrificioBienestarSi']
        ws['F49'] = datos['sacrificioBienestarNo']
        ws['G49'] = datos['sacrificioBienestarNa']

        # Otros campos
        ws['B52'] = datos['na']
        ws['B54'] = datos['observaciones']

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_file_path = temp_file.name
            wb.save(temp_file_path)
            print("Archivo modificado con exito")

        # Verifica que el archivo se guardó correctamente
        if not os.path.exists(temp_file_path):
            return jsonify({"status": "error", "message": "El archivo no se pudo generar."}), 500

        # Devuelve el archivo para su descarga
        
        print("funciona")
        return send_file(temp_file_path, as_attachment=True)




if __name__ == '__main__':
    app.run(debug=True)


